#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
引物订购表数据提取脚本
从生工生物的引物合成订购表中提取订单信息和引物数据
"""

import openpyxl
from typing import Dict, List, Any
import json
import sys
import shutil
from pathlib import Path
from datetime import datetime


class PrimerOrderExtractor:
    """引物订购表数据提取器"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet = None
        
        # 订单信息字段映射
        self.field_mapping = {
            "订购人姓名": "customer_name",
            "订购人手机": "customer_phone",
            "订购人E-MAIL": "customer_email",
            "负责人/法人": "responsible_person",
            "固定电话": "company_phone",
            "订购人单位": "company_name",
            "发票抬头": "invoice_title",
            "付款方式": "payment_method",
            "收货地址": "delivery_address",
            "备注": "notes",
            "订购公司": "ordering_company",
            "订购日期": "order_date",
        }
        
        # 布尔字段映射
        self.boolean_fields = {
            "随货开票": "invoice_with_goods",
            "是否双休日发货": "weekend_delivery",
            "是否部分先发货": "partial_delivery",
        }
        
        # 引物数据字段映射
        self.primer_field_mapping = {
            "Primer名称": "primer_name",
            "序列（5' to 3'）": "sequence",
            "序列": "sequence",
            "碱基数": "base_count",
            "分装管数": "tube_count",
            "提供总量（OD）": "total_quantity",
            "提供总量": "total_quantity",
            "纯化方法": "purification_method",
            "nmoles": "nmoles",
            "5'端修饰": "five_modification",
            "3'端修饰": "three_modification",
            "5'端和3'端双标记修饰": "double_modification",
            "类型": "primer_type",
            "备注": "remarks",
        }
    
    def load_excel(self):
        """加载Excel文件"""
        try:
            # 使用data_only=True来读取公式的计算值而不是公式本身
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            self.worksheet = self.workbook.active
            print(f"成功加载Excel文件: {self.excel_path}")
        except Exception as e:
            print(f"加载Excel文件失败: {e}")
            sys.exit(1)
    
    def _safe_cell_value(self, cell_value):
        """安全地提取单元格值，处理公式对象等特殊情况"""
        if cell_value is None:
            return None
        
        # 处理公式对象 - 递归提取值
        if hasattr(cell_value, 'value'):
            return self._safe_cell_value(cell_value.value)
        
        # 处理datetime对象
        if hasattr(cell_value, 'strftime'):
            # 是datetime对象，转换为字符串
            try:
                return cell_value.strftime('%Y/%m/%d')
            except:
                return str(cell_value)
        
        # 处理基本类型
        if isinstance(cell_value, (int, float, str, bool)):
            return cell_value
        
        # 对于无法识别的类型返回None而不是字符串
        # 这样可以避免将对象表示形式作为数据
        return None
    
    def extract_order_info(self) -> Dict[str, Any]:
        """提取订单信息（前面几行的信息）"""
        order_info = {}
        
        # 特殊处理：提取右上角的订购日期
        # 扫描前5行的右侧列 (J到O列，即10-15列)
        for row in range(1, 6):
            for col in range(10, 16):  # J到O列
                cell_value = self._safe_cell_value(self.worksheet.cell(row, col).value)
                if cell_value and isinstance(cell_value, str) and "订购日期" in cell_value:
                    # 找到"订购日期"标签
                    # 检查标签后面是否直接跟着冒号和日期
                    if ":" in cell_value or "：" in cell_value:
                        # 日期可能在同一单元格中
                        parts = cell_value.replace("：", ":").split(":")
                        if len(parts) > 1:
                            date_str = parts[1].strip()
                            if date_str:
                                order_info['order_date'] = date_str
                                break
                    
                    # 尝试从右侧单元格获取
                    date_value = self._safe_cell_value(self.worksheet.cell(row, col + 1).value)
                    if date_value:
                        order_info['order_date'] = str(date_value).strip()
                        break
                    
                    # 尝试从下一行同列获取
                    date_value = self._safe_cell_value(self.worksheet.cell(row + 1, col).value)
                    if date_value:
                        order_info['order_date'] = str(date_value).strip()
                        break
            
            if 'order_date' in order_info:
                break
        
        # 如果还没找到，尝试直接扫描右上角区域寻找日期格式的值
        if 'order_date' not in order_info:
            for row in range(1, 6):
                for col in range(12, 16):  # M到O列
                    cell_value = self._safe_cell_value(self.worksheet.cell(row, col).value)
                    if cell_value:
                        value_str = str(cell_value).strip()
                        # 检查是否像日期 (包含/或-，或纯数字日期格式)
                        if ('/' in value_str or '-' in value_str) and len(value_str) >= 8:
                            order_info['order_date'] = value_str
                            break
                        # 或者是YYYY/MM/DD格式
                        if len(value_str) == 10 and (value_str[4] == '/' or value_str[4] == '-'):
                            order_info['order_date'] = value_str
                            break
                if 'order_date' in order_info:
                    break
        
        # 遍历前30行查找其他订单信息
        for row in range(1, 31):
            for col in range(1, 10):
                cell_value = self._safe_cell_value(self.worksheet.cell(row, col).value)
                
                if cell_value and isinstance(cell_value, str):
                    # 检查普通字段
                    for chinese_key, english_key in self.field_mapping.items():
                        if chinese_key in cell_value:
                            # 尝试从同一行右侧单元格获取值
                            value = self._safe_cell_value(self.worksheet.cell(row, col + 1).value)
                            if value:
                                order_info[english_key] = str(value).strip()
                            break
                    
                    # 检查布尔字段
                    for chinese_key, english_key in self.boolean_fields.items():
                        if chinese_key in cell_value:
                            # 查找"是"字样
                            value = False
                            for check_col in range(col, col + 5):
                                check_value = self._safe_cell_value(self.worksheet.cell(row, check_col).value)
                                if check_value and "是" in str(check_value):
                                    value = True
                                    break
                            order_info[english_key] = value
                            break
        
        return order_info
    
    def find_primer_table_start(self) -> int:
        """查找引物数据表的起始行（包含表头的行）"""
        for row in range(1, 50):
            for col in range(1, 20):
                cell_value = self._safe_cell_value(self.worksheet.cell(row, col).value)
                if cell_value and isinstance(cell_value, str):
                    # 查找"Primer名称"或"序列"等关键字
                    if "Primer名称" in cell_value or "序列（5' to 3'）" in cell_value:
                        return row
        return None
    
    def extract_primer_data(self) -> List[Dict[str, Any]]:
        """提取引物数据"""
        primer_data = []
        
        # 查找引物表起始行
        header_row = self.find_primer_table_start()
        if not header_row:
            print("未找到引物数据表")
            return primer_data
        
        print(f"找到引物数据表，表头在第 {header_row} 行")
        
        # 读取表头
        headers = []
        header_col_map = {}  # 列索引到英文字段名的映射
        
        for col in range(1, 30):
            header_value = self._safe_cell_value(self.worksheet.cell(header_row, col).value)
            if header_value:
                header_str = str(header_value).strip()
                headers.append(header_str)
                
                # 映射到英文字段名
                for chinese_key, english_key in self.primer_field_mapping.items():
                    if chinese_key in header_str:
                        header_col_map[col] = english_key
                        break
            else:
                headers.append(None)
        
        print(f"识别到的列: {header_col_map}")
        
        # 读取数据行
        data_row = header_row + 1
        consecutive_empty_rows = 0
        max_empty_rows = 10  # 连续10行空行才停止
        
        while data_row <= self.worksheet.max_row:
            # 提取该行数据
            row_data = {}
            has_data = False
            
            for col, english_key in header_col_map.items():
                cell = self.worksheet.cell(data_row, col)
                safe_value = self._safe_cell_value(cell.value)
                
                if safe_value is not None:
                    row_data[english_key] = safe_value
                    has_data = True
            
            if has_data:
                # 检查是否有primer_name，没有则跳过此行
                if 'primer_name' not in row_data or not row_data['primer_name']:
                    # primer_name为空，跳过此行
                    consecutive_empty_rows += 1
                    if consecutive_empty_rows >= max_empty_rows:
                        print(f"在第 {data_row} 行遇到连续 {max_empty_rows} 个空行，停止解析")
                        break
                else:
                    # 后处理：如果base_count为空但有sequence，从sequence计算
                    if ('base_count' not in row_data or row_data['base_count'] is None) and 'sequence' in row_data:
                        sequence = row_data['sequence']
                        if sequence and isinstance(sequence, str):
                            row_data['base_count'] = len(sequence.strip())
                    
                    # 添加行号
                    row_data['row_number'] = data_row
                    primer_data.append(row_data)
                    consecutive_empty_rows = 0  # 重置空行计数
            else:
                consecutive_empty_rows += 1
                # 连续空行超过阈值则停止
                if consecutive_empty_rows >= max_empty_rows:
                    print(f"在第 {data_row} 行遇到连续 {max_empty_rows} 个空行，停止解析")
                    break
            
            data_row += 1
        
        print(f"提取到 {len(primer_data)} 条引物数据")
        return primer_data
    
    def extract_all(self) -> Dict[str, Any]:
        """提取所有数据"""
        self.load_excel()
        
        result = {
            "order_info": self.extract_order_info(),
            "primer_data": self.extract_primer_data()
        }
        
        return result
    
    def save_to_json(self, output_path: str):
        """保存为JSON文件"""
        data = self.extract_all()

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"\n数据已保存到: {output_path}")
        return data

    def save_to_db(self, data: Dict[str, Any] = None) -> int:
        """
        将提取结果存入 skita 数据库。

        1. 安装 schema 到项目 meta/
        2. ensure_table 建表
        3. insert 记录

        Returns:
            插入记录的 id
        """
        if data is None:
            data = self.extract_all()

        # --- 导入项目工具 ---
        root = Path(__file__).resolve().parent.parent.parent.parent.parent
        sys.path.insert(0, str(root))
        from scripts.db import DB
        from scripts.files import Files

        # --- 安装 schema ---
        skill_meta = Path(__file__).resolve().parent.parent / "meta"
        project_meta = root / "meta"
        project_meta.mkdir(parents=True, exist_ok=True)
        schema_file = skill_meta / "sangon_primer_order.json"
        shutil.copy2(str(schema_file), str(project_meta / schema_file.name))

        # --- 建表 ---
        db = DB()
        db.ensure_table("sangon_primer_order")

        # --- 归档源文件 ---
        source_rel = None
        excel_abs = Path(self.excel_path).resolve()
        if excel_abs.exists():
            result = Files().save(str(excel_abs), category="data-preprocess")
            source_rel = result["relative_path"]

        # --- 组装记录 ---
        order_info = data.get("order_info", {})
        primer_data = data.get("primer_data", [])
        record = {
            "source_file": source_rel or str(excel_abs),
            "customer_name": order_info.get("customer_name"),
            "customer_phone": order_info.get("customer_phone"),
            "customer_email": order_info.get("customer_email"),
            "company_name": order_info.get("company_name"),
            "order_date": order_info.get("order_date"),
            "payment_method": order_info.get("payment_method"),
            "primer_count": len(primer_data),
            "primer_data": json.dumps(primer_data, ensure_ascii=False),
            "order_info": json.dumps(order_info, ensure_ascii=False),
        }

        row_id = db.insert("data_sangon_primer_order", record)
        print(f"已存入数据库，记录 id={row_id}，共 {len(primer_data)} 条引物")
        return row_id

    def print_summary(self, data: Dict[str, Any]):
        """打印数据摘要"""
        print("\n" + "="*60)
        print("订单信息摘要")
        print("="*60)

        order_info = data.get("order_info", {})
        for key, value in order_info.items():
            print(f"{key}: {value}")

        print("\n" + "="*60)
        print("引物数据摘要")
        print("="*60)

        primer_data = data.get("primer_data", [])
        print(f"共 {len(primer_data)} 条引物数据")

        if primer_data:
            print("\n前3条引物数据预览:")
            for i, primer in enumerate(primer_data[:3], 1):
                print(f"\n引物 {i}:")
                for key, value in primer.items():
                    print(f"  {key}: {value}")


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("使用方法: python extract_sangon_order_xlsx.py <excel文件路径> [--save-db] [--json output.json]")
        print("示例:")
        print("  python extract_sangon_order_xlsx.py order.xlsx                  # 仅提取并打印")
        print("  python extract_sangon_order_xlsx.py order.xlsx --save-db        # 提取并存入数据库")
        print("  python extract_sangon_order_xlsx.py order.xlsx --json out.json  # 提取并保存为JSON")
        sys.exit(1)

    excel_path = sys.argv[1]
    args = sys.argv[2:]

    save_db = "--save-db" in args
    json_path = None
    if "--json" in args:
        idx = args.index("--json")
        if idx + 1 < len(args):
            json_path = args[idx + 1]

    extractor = PrimerOrderExtractor(excel_path)
    data = extractor.extract_all()

    if json_path:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"JSON 已保存到: {json_path}")

    if save_db:
        extractor.save_to_db(data)

    extractor.print_summary(data)


if __name__ == "__main__":
    main()